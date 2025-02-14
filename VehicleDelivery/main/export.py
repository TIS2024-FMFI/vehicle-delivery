import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.db.models import FileField, ImageField, DateField, Count
from django.utils.timezone import now
from datetime import timedelta
from main.models import *
import zipfile
import os
from main.logging import get_complaint_type


def export_single_object(request, obj_id, model):
    obj = get_object_or_404(model, id=obj_id)

    # Create a new workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Complaint Data"

    # Get all field names, excluding FileField and ImageField
    fields = [field for field in model._meta.fields if not isinstance(field, (FileField, ImageField))]
    headers = [field.name for field in fields]
    worksheet.append(headers)

    # Apply styling to headers
    header_fill = PatternFill(start_color="B3D9FF", end_color="B3D9FF", fill_type="solid")
    for col_num, header_cell in enumerate(worksheet[1], 1):
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="center")
        header_cell.fill = header_fill

    values = []
    for field in fields:
        value = getattr(obj, field.name)

        # If the field is a DateField, convert it to a string for correct formatting
        if isinstance(field, DateField) and value:
            value = value.strftime("%Y-%m-%d")  # Format dates

        values.append(value)

    worksheet.append(values)

    # Auto-adjust column widths with a max limit
    MAX_COLUMN_WIDTH = 25  # Set a max width (characters)
    for col_num, col_cells in enumerate(worksheet.columns, 1):
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col_num) 
        for cell in col_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, MAX_COLUMN_WIDTH)  # Apply max width limit
        worksheet.column_dimensions[col_letter].width = adjusted_width

    # Prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="complaint_{obj_id}.xlsx"'

    workbook.save(response)

    ActionLog(
        user=request.user.person,
        target_type=get_complaint_type(model),
        target_id=obj_id,
        action="export",
    ).save()
    

    return response



def download_all_files(request, obj_id, model):
    obj = get_object_or_404(model, id=obj_id)

    # Collect all file fields
    file_fields = [
        field.name for field in obj._meta.fields 
        if isinstance(field, (models.FileField, models.ImageField))
    ]
    
    # Collect existing files
    files = [getattr(obj, field) for field in file_fields if getattr(obj, field)]

    # If no files, return an error response
    if not files:
        return HttpResponse("No files available for download.", content_type="text/plain")

    # Create ZIP file
    response = HttpResponse(content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="documents_{obj_id}.zip"'

    with zipfile.ZipFile(response, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for file in files:
            file_path = file.path 
            if os.path.exists(file_path):
                zip_file.write(file_path, f"{obj_id}_{os.path.basename(file_path)}")

    ActionLog(
        user=request.user.person,
        target_type=get_complaint_type(model),
        target_id=obj_id,
        action="download files",
    ).save()

    return response


def export_statistics(request, context):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Statistics"

    # A function to write table data to an Excel sheet with formatted headers and auto-adjusted column widths.
    def add_table(sheet, row, title, columns, data):
        
        sheet.cell(row=row, column=1, value=title)  # Title row
        row += 1  # Move to the header row

        # Apply header styles
        header_fill = PatternFill(start_color="B3D9FF", end_color="B3D9FF", fill_type="solid")  # Light Blue
        header_font = Font(bold=True)

        #sheet.append(columns)  # Column headers
        for col_num, col_name in enumerate(columns, 1):  # 1-based index
            cell = sheet.cell(row=row, column=col_num, value=col_name)
            cell.fill = header_fill
            cell.font = header_font

        row += 1  # Move to data row

        # Write table data
        for entry in data:
            row_values = [str(entry.get(col_name, "")) for col_name in columns]  # Get values safely
            sheet.append(row_values)

        # Adjust column widths based on content
        for col_num, col_name in enumerate(columns, 1):
            column_letter = get_column_letter(col_num)
            max_length = max(len(str(entry.get(col_name, ""))) for entry in data) if data else len(col_name)
            adjusted_width = max_length + 4  # Add padding
            sheet.column_dimensions[column_letter].width = max(adjusted_width, 13)

        return row + len(data) + 2  # Return updated row position

    
    sheet.append(["start date", "end date"])
    sheet.append([context["start_date"], context["end_date"]])
    row = 5
    # Imports per Month Table
    row = add_table(sheet, row, "Complaints Imported", ["target_type", "count", "percentage"], context["imports_per_month"])

    # Status Changes Table
    row = add_table(sheet, row, "Status Changes", ["target_type", "new_value", "count", "percentage"], context["status_changes"])

    # Nature of Damage Table
    row = add_table(sheet, row, "Nature of Damage", ["code", "name", "count", "percentage"], context["nature_of_damage_counts"])

    # Place of Damage Table
    row = add_table(sheet, row, "Place of Damage", ["code", "name", "count", "percentage"], context["place_of_damage_counts"])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="statistics_{context["start_date"]}_{context["end_date"]}.xlsx"'

    workbook.save(response)
    return response
